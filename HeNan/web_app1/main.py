from fastapi import FastAPI, HTTPException, Query, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import pymysql
import pandas as pd
import json
import os
import re
from io import BytesIO
from typing import Optional, List, Dict, Any
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from pydantic import BaseModel
from urllib.parse import quote
from pypinyin import pinyin, Style

app = FastAPI(title="运营管理平台", description="剧集信息管理系统")

# 配置CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 数据库连接配置
DB_CONFIG = {
    'host': 'localhost',
    'port': 3306,
    'user': 'root',
    'password': 'polarbear',
    'database': 'operation_management',
    'charset': 'utf8mb4'
}

# 剧头数据列名顺序
DRAMA_HEADER_COLUMNS = [
    '剧头id', '剧集名称', '作者列表', '清晰度', '语言', '主演', '内容类型', '上映年份',
    '关键字', '评分', '推荐语', '总集数', '产品分类', '竖图', '描述', '横图', '版权', '二级分类'
]

# 子集数据列名顺序
SUBSET_COLUMNS = [
    '子集id', '节目名称', '媒体拉取地址', '媒体类型', '编码格式', '集数', '时长', '文件大小'
]

# 静态文件服务
import os
from pathlib import Path

BASE_DIR = Path(__file__).parent
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")


def get_db_connection():
    """获取数据库连接"""
    return pymysql.connect(**DB_CONFIG)


@app.get("/")
async def read_root():
    """返回首页"""
    return FileResponse(str(BASE_DIR / "index.html"))


@app.get("/api/customers")
async def get_customers():
    """获取客户列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 查询所有客户
        query = """
            SELECT customer_id, customer_name, customer_code, remark, created_at, updated_at
            FROM customer
            ORDER BY created_at DESC
        """
        cursor.execute(query)
        customers = cursor.fetchall()
        
        # 为每个客户查询剧集数量
        for customer in customers:
            count_query = "SELECT COUNT(*) as count FROM drama_main WHERE customer_id = %s"
            cursor.execute(count_query, (customer['customer_id'],))
            result = cursor.fetchone()
            customer['drama_count'] = result['count'] if result else 0
        
        conn.close()
        
        return {
            "code": 200,
            "message": "success",
            "data": customers
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dramas")
async def get_dramas(
    customer_id: Optional[int] = Query(None, description="客户ID"),
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(10, ge=1, le=100, description="每页数量")
):
    """获取剧集列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 构建查询条件
        where_conditions = []
        params = []
        
        if customer_id:
            where_conditions.append("customer_id = %s")
            params.append(customer_id)
        
        if keyword:
            where_conditions.append("drama_name LIKE %s")
            params.append(f"%{keyword}%")
        
        where_clause = ""
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions)
        
        # 查询总数
        count_query = f"SELECT COUNT(*) as total FROM drama_main {where_clause}"
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        
        # 查询数据
        offset = (page - 1) * page_size
        query = f"""
            SELECT drama_id, customer_id, drama_name, dynamic_properties, created_at, updated_at
            FROM drama_main
            {where_clause}
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
        """
        cursor.execute(query, params + [page_size, offset])
        dramas = cursor.fetchall()
        
        # 解析dynamic_properties
        for drama in dramas:
            if drama['dynamic_properties']:
                if isinstance(drama['dynamic_properties'], str):
                    drama['dynamic_properties'] = json.loads(drama['dynamic_properties'])
        
        conn.close()
        
        return {
            "code": 200,
            "message": "success",
            "data": {
                "list": dramas,
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dramas/by-name")
async def get_drama_by_name(name: str = Query(..., description="剧集名称")):
    """根据剧集名称获取剧集详情（包含子集）"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 查询剧头数据
        query = "SELECT * FROM drama_main WHERE drama_name = %s"
        cursor.execute(query, (name,))
        drama = cursor.fetchone()
        
        if not drama:
            conn.close()
            raise HTTPException(status_code=404, detail="未找到该剧集")
        
        # 解析dynamic_properties
        if drama['dynamic_properties']:
            if isinstance(drama['dynamic_properties'], str):
                drama['dynamic_properties'] = json.loads(drama['dynamic_properties'])
        
        # 构建完整的剧头数据
        dynamic_props = drama['dynamic_properties'] or {}
        header_dict = {
            '剧头id': drama['drama_id'],
            '剧集名称': drama['drama_name'],
            '作者列表': dynamic_props.get('作者列表', ''),
            '清晰度': dynamic_props.get('清晰度', 0),
            '语言': dynamic_props.get('语言', ''),
            '主演': dynamic_props.get('主演', ''),
            '内容类型': dynamic_props.get('内容类型', ''),
            '上映年份': dynamic_props.get('上映年份', 0),
            '关键字': dynamic_props.get('关键字', ''),
            '评分': dynamic_props.get('评分', 0.0),
            '推荐语': dynamic_props.get('推荐语', ''),
            '总集数': dynamic_props.get('总集数', 0),
            '产品分类': dynamic_props.get('产品分类', 0),
            '竖图': dynamic_props.get('竖图', ''),
            '描述': dynamic_props.get('描述', ''),
            '横图': dynamic_props.get('横图', ''),
            '版权': dynamic_props.get('版权', 0),
            '二级分类': dynamic_props.get('二级分类', '')
        }
        
        # 查询子集数据
        query_episodes = "SELECT * FROM drama_episode WHERE drama_id = %s ORDER BY episode_id"
        cursor.execute(query_episodes, (drama['drama_id'],))
        episodes = cursor.fetchall()
        
        episode_list = []
        for i, episode in enumerate(episodes, 1):
            dynamic_props_ep = {}
            if episode['dynamic_properties']:
                if isinstance(episode['dynamic_properties'], str):
                    dynamic_props_ep = json.loads(episode['dynamic_properties'])
            
            episode_data = {
                '子集id': i,
                '节目名称': episode['episode_name'],
                '媒体拉取地址': dynamic_props_ep.get('媒体拉取地址', ''),
                '媒体类型': dynamic_props_ep.get('媒体类型', 0),
                '编码格式': dynamic_props_ep.get('编码格式', 0),
                '集数': dynamic_props_ep.get('集数', 0),
                '时长': dynamic_props_ep.get('时长', 0),
                '文件大小': dynamic_props_ep.get('文件大小', 0)
            }
            episode_list.append(episode_data)
        
        conn.close()
        
        return {
            "code": 200,
            "message": "success",
            "data": {
                "header": header_dict,
                "episodes": episode_list
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dramas/{drama_id}")
async def get_drama_detail(drama_id: int):
    """获取剧集详细信息"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 查询剧头数据
        query = "SELECT * FROM drama_main WHERE drama_id = %s"
        cursor.execute(query, (drama_id,))
        drama = cursor.fetchone()
        
        if not drama:
            conn.close()
            raise HTTPException(status_code=404, detail="剧集不存在")
        
        # 解析dynamic_properties
        if drama['dynamic_properties']:
            if isinstance(drama['dynamic_properties'], str):
                drama['dynamic_properties'] = json.loads(drama['dynamic_properties'])
        
        # 构建完整的剧头数据
        dynamic_props = drama['dynamic_properties'] or {}
        header_dict = {
            '剧头id': drama['drama_id'],
            '剧集名称': drama['drama_name'],
            '作者列表': dynamic_props.get('作者列表', ''),
            '清晰度': dynamic_props.get('清晰度', 0),
            '语言': dynamic_props.get('语言', ''),
            '主演': dynamic_props.get('主演', ''),
            '内容类型': dynamic_props.get('内容类型', ''),
            '上映年份': dynamic_props.get('上映年份', 0),
            '关键字': dynamic_props.get('关键字', ''),
            '评分': dynamic_props.get('评分', 0.0),
            '推荐语': dynamic_props.get('推荐语', ''),
            '总集数': dynamic_props.get('总集数', 0),
            '产品分类': dynamic_props.get('产品分类', 0),
            '竖图': dynamic_props.get('竖图', ''),
            '描述': dynamic_props.get('描述', ''),
            '横图': dynamic_props.get('横图', ''),
            '版权': dynamic_props.get('版权', 0),
            '二级分类': dynamic_props.get('二级分类', '')
        }
        
        conn.close()
        
        return {
            "code": 200,
            "message": "success",
            "data": header_dict
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dramas/{drama_id}/episodes")
async def get_drama_episodes(drama_id: int):
    """获取剧集的子集列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 查询子集数据
        query = "SELECT * FROM drama_episode WHERE drama_id = %s ORDER BY episode_id"
        cursor.execute(query, (drama_id,))
        episodes = cursor.fetchall()
        
        # 解析dynamic_properties并构建子集数据
        episode_list = []
        for i, episode in enumerate(episodes, 1):
            dynamic_props = {}
            if episode['dynamic_properties']:
                if isinstance(episode['dynamic_properties'], str):
                    dynamic_props = json.loads(episode['dynamic_properties'])
            
            episode_data = {
                '子集id': episode['episode_id'],
                '节目名称': episode['episode_name'],
                '媒体拉取地址': dynamic_props.get('媒体拉取地址', ''),
                '媒体类型': dynamic_props.get('媒体类型', 0),
                '编码格式': dynamic_props.get('编码格式', 0),
                '集数': dynamic_props.get('集数', 0),
                '时长': dynamic_props.get('时长', 0),
                '文件大小': dynamic_props.get('文件大小', 0)
            }
            episode_list.append(episode_data)
        
        conn.close()
        
        return {
            "code": 200,
            "message": "success",
            "data": episode_list
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/dramas")
async def create_drama(drama_data: Dict[str, Any] = Body(...)):
    """创建剧头"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 检查必填字段
        if '剧集名称' not in drama_data or not drama_data['剧集名称']:
            conn.close()
            raise HTTPException(status_code=400, detail="剧集名称不能为空")
        
        drama_name = drama_data['剧集名称']
        customer_id = drama_data.get('customer_id')
        
        # 构建dynamic_properties字典
        dynamic_props = {}
        
        # 需要存储到dynamic_properties中的字段
        dynamic_fields = [
            '作者列表', '清晰度', '语言', '主演', '内容类型', '上映年份',
            '关键字', '评分', '推荐语', '总集数', '产品分类', '竖图',
            '描述', '横图', '版权', '二级分类'
        ]
        
        for field in dynamic_fields:
            if field in drama_data:
                dynamic_props[field] = drama_data[field]
        
        # 插入数据
        insert_query = "INSERT INTO drama_main (customer_id, drama_name, dynamic_properties) VALUES (%s, %s, %s)"
        dynamic_props_json = json.dumps(dynamic_props, ensure_ascii=False) if dynamic_props else None
        cursor.execute(insert_query, (customer_id, drama_name, dynamic_props_json))
        conn.commit()
        
        new_drama_id = cursor.lastrowid
        conn.close()
        
        return {
            "code": 200,
            "message": "创建成功",
            "data": {"drama_id": new_drama_id}
        }
    except HTTPException:
        if conn:
            conn.close()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/dramas/{drama_id}")
async def delete_drama(drama_id: int):
    """删除剧头"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 检查剧集是否存在
        check_query = "SELECT * FROM drama_main WHERE drama_id = %s"
        cursor.execute(check_query, (drama_id,))
        drama = cursor.fetchone()
        
        if not drama:
            conn.close()
            raise HTTPException(status_code=404, detail="剧集不存在")
        
        # 删除剧集（外键约束会自动删除关联的子集）
        delete_query = "DELETE FROM drama_main WHERE drama_id = %s"
        cursor.execute(delete_query, (drama_id,))
        conn.commit()
        
        conn.close()
        
        return {
            "code": 200,
            "message": "删除成功",
            "data": {"drama_id": drama_id}
        }
    except HTTPException:
        if conn:
            conn.close()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/dramas/{drama_id}")
async def update_drama(drama_id: int, drama_data: Dict[str, Any] = Body(...)):
    """更新剧集信息"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 检查剧集是否存在
        check_query = "SELECT * FROM drama_main WHERE drama_id = %s"
        cursor.execute(check_query, (drama_id,))
        drama = cursor.fetchone()
        
        if not drama:
            conn.close()
            raise HTTPException(status_code=404, detail="剧集不存在")
        
        # 构建dynamic_properties字典
        dynamic_props = {}
        
        # 需要存储到dynamic_properties中的字段
        dynamic_fields = [
            '作者列表', '清晰度', '语言', '主演', '内容类型', '上映年份',
            '关键字', '评分', '推荐语', '总集数', '产品分类', '竖图',
            '描述', '横图', '版权', '二级分类'
        ]
        
        for field in dynamic_fields:
            if field in drama_data:
                dynamic_props[field] = drama_data[field]
        
        # 更新drama_name（如果提供）
        update_fields = []
        update_values = []
        
        if '剧集名称' in drama_data:
            update_fields.append("drama_name = %s")
            update_values.append(drama_data['剧集名称'])
        
        # 更新dynamic_properties
        if dynamic_props:
            update_fields.append("dynamic_properties = %s")
            update_values.append(json.dumps(dynamic_props, ensure_ascii=False))
        
        if not update_fields:
            conn.close()
            return {
                "code": 400,
                "message": "没有提供要更新的字段",
                "data": None
            }
        
        # 执行更新
        update_query = f"UPDATE drama_main SET {', '.join(update_fields)} WHERE drama_id = %s"
        update_values.append(drama_id)
        cursor.execute(update_query, update_values)
        conn.commit()
        
        conn.close()
        
        return {
            "code": 200,
            "message": "更新成功",
            "data": {"drama_id": drama_id}
        }
    except HTTPException:
        if conn:
            conn.close()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/dramas/{drama_id}/export")
async def export_drama_to_excel(drama_id: int):
    """导出剧集数据为Excel文件"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 查询剧头数据
        query = "SELECT * FROM drama_main WHERE drama_id = %s"
        cursor.execute(query, (drama_id,))
        drama = cursor.fetchone()
        
        if not drama:
            conn.close()
            raise HTTPException(status_code=404, detail="剧集不存在")
        
        drama_name = drama['drama_name']
        
        # 解析dynamic_properties
        dynamic_props = {}
        if drama['dynamic_properties']:
            if isinstance(drama['dynamic_properties'], str):
                dynamic_props = json.loads(drama['dynamic_properties'])
        
        # 构建剧头数据
        header_dict = {
            '剧头id': '',
            '剧集名称': drama['drama_name'],
            '作者列表': dynamic_props.get('作者列表', ''),
            '清晰度': dynamic_props.get('清晰度', 0),
            '语言': dynamic_props.get('语言', ''),
            '主演': dynamic_props.get('主演', ''),
            '内容类型': dynamic_props.get('内容类型', ''),
            '上映年份': dynamic_props.get('上映年份', 0),
            '关键字': dynamic_props.get('关键字', ''),
            '评分': dynamic_props.get('评分', 0.0),
            '推荐语': dynamic_props.get('推荐语', ''),
            '总集数': dynamic_props.get('总集数', 0),
            '产品分类': dynamic_props.get('产品分类', 0),
            '竖图': dynamic_props.get('竖图', ''),
            '描述': dynamic_props.get('描述', ''),
            '横图': dynamic_props.get('横图', ''),
            '版权': dynamic_props.get('版权', 0),
            '二级分类': dynamic_props.get('二级分类', '')
        }
        header_df = pd.DataFrame([header_dict], columns=DRAMA_HEADER_COLUMNS)
        
        # 查询子集数据
        query_episodes = "SELECT * FROM drama_episode WHERE drama_id = %s ORDER BY episode_id"
        cursor.execute(query_episodes, (drama['drama_id'],))
        episodes = cursor.fetchall()
        
        episode_list = []
        for i, episode in enumerate(episodes, 1):
            dynamic_props_ep = {}
            if episode['dynamic_properties']:
                if isinstance(episode['dynamic_properties'], str):
                    dynamic_props_ep = json.loads(episode['dynamic_properties'])
            
            episode_data = {
                '子集id': '',
                '节目名称': episode['episode_name'],
                '媒体拉取地址': dynamic_props_ep.get('媒体拉取地址', ''),
                '媒体类型': dynamic_props_ep.get('媒体类型', 0),
                '编码格式': dynamic_props_ep.get('编码格式', 0),
                '集数': dynamic_props_ep.get('集数', 0),
                '时长': dynamic_props_ep.get('时长', 0),
                '文件大小': dynamic_props_ep.get('文件大小', 0)
            }
            episode_list.append(episode_data)
        
        subset_df = pd.DataFrame(episode_list, columns=SUBSET_COLUMNS)
        
        conn.close()
        
        # 生成Excel文件到内存
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            header_df.to_excel(writer, sheet_name='剧头', index=False)
            subset_df.to_excel(writer, sheet_name='子集', index=False)
            
            # 获取workbook和worksheets
            workbook = writer.book
            
            # 调整剧头工作表的列宽和文本换行
            header_sheet = workbook['剧头']
            header_sheet.row_dimensions[1].height = 30
            for row_idx in range(2, len(header_df) + 2):
                header_sheet.row_dimensions[row_idx].height = None
            
            for idx, col in enumerate(header_df.columns, 1):
                column_letter = get_column_letter(idx)
                col_name = str(col)
                col_name_width = sum(2 if ord(c) > 127 else 1 for c in col_name)
                max_data_width = col_name_width
                for row_idx, value in enumerate(header_df[col], 2):
                    if value is not None:
                        value_str = str(value)
                        data_width = sum(2 if ord(c) > 127 else 1 for c in value_str)
                        max_data_width = max(max_data_width, data_width)
                        cell = header_sheet.cell(row=row_idx, column=idx)
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                header_cell = header_sheet.cell(row=1, column=idx)
                header_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                column_width = max(col_name_width, max_data_width) * 1.3 + 3
                header_sheet.column_dimensions[column_letter].width = column_width
            
            # 调整子集工作表的列宽和文本换行
            subset_sheet = workbook['子集']
            subset_sheet.row_dimensions[1].height = 30
            for row_idx in range(2, len(subset_df) + 2):
                subset_sheet.row_dimensions[row_idx].height = None
            
            for idx, col in enumerate(subset_df.columns, 1):
                column_letter = get_column_letter(idx)
                col_name = str(col)
                col_name_width = sum(2 if ord(c) > 127 else 1 for c in col_name)
                max_data_width = col_name_width
                for row_idx, value in enumerate(subset_df[col], 2):
                    if value is not None:
                        value_str = str(value)
                        data_width = sum(2 if ord(c) > 127 else 1 for c in value_str)
                        max_data_width = max(max_data_width, data_width)
                        cell = subset_sheet.cell(row=row_idx, column=idx)
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                header_cell = subset_sheet.cell(row=1, column=idx)
                header_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                column_width = max(col_name_width, max_data_width) * 1.3 + 3
                subset_sheet.column_dimensions[column_letter].width = column_width
        
        output.seek(0)
        
        # 对中文文件名进行编码
        encoded_filename = quote(f"{drama_name}_数据.xlsx")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 版权方数据 API ====================

@app.get("/api/copyright")
async def get_copyright_list(
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(10, ge=1, le=100, description="每页数量")
):
    """获取版权方数据列表"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        where_conditions = []
        params = []
        
        if keyword:
            where_conditions.append("media_name LIKE %s")
            params.append(f"%{keyword}%")
        
        where_clause = ""
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions)
        
        # 查询总数
        count_query = f"SELECT COUNT(*) as total FROM copyright_content {where_clause}"
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']
        
        # 查询数据
        offset = (page - 1) * page_size
        query = f"""
            SELECT * FROM copyright_content
            {where_clause}
            ORDER BY id DESC
            LIMIT %s OFFSET %s
        """
        cursor.execute(query, params + [page_size, offset])
        items = cursor.fetchall()
        
        conn.close()
        
        return {
            "code": 200,
            "message": "success",
            "data": {
                "list": items,
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/copyright/{item_id}")
async def get_copyright_detail(item_id: int):
    """获取版权方数据详情"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        query = "SELECT * FROM copyright_content WHERE id = %s"
        cursor.execute(query, (item_id,))
        item = cursor.fetchone()
        
        conn.close()
        
        if not item:
            raise HTTPException(status_code=404, detail="数据不存在")
        
        return {
            "code": 200,
            "message": "success",
            "data": item
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/copyright")
async def create_copyright(data: Dict[str, Any] = Body(...)):
    """创建版权方数据，并自动创建对应的剧头和子集"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        if 'media_name' not in data or not data['media_name']:
            raise HTTPException(status_code=400, detail="介质名称不能为空")
        
        media_name = data['media_name']
        customer_id = data.get('customer_id')  # 获取客户ID
        
        # 1. 先创建剧头数据
        # 字段映射
        author_list = data.get('director') or "暂无"
        language = data.get('language_henan') or data.get('language') or "简体中文"
        actors = data.get('cast_members') or ""
        content_type = data.get('category_level1_henan') or "电视剧"
        release_year = int(data['production_year']) if data.get('production_year') else None
        keywords = data.get('keywords') or ""
        rating = float(data['rating']) if data.get('rating') else None
        recommendation = data.get('recommendation') or ""
        total_episodes = int(data['episode_count']) if data.get('episode_count') else 0
        description = data.get('synopsis') or ""
        secondary_category = data.get('category_level2_henan') or ""
        
        # 产品分类：教育=1, 电竞=2, 其他=3
        product_category = 3
        if content_type:
            if "教育" in str(content_type):
                product_category = 1
            elif "电竞" in str(content_type):
                product_category = 2
        
        # 生成图片URL（使用剧集名称每个字的汉语拼音首字母）
        def get_pinyin_abbr(name):
            result = []
            for char in name:
                if '\u4e00' <= char <= '\u9fff':  # 中文字符
                    py = pinyin(char, style=Style.FIRST_LETTER)
                    if py and py[0]:
                        result.append(py[0][0])
                elif char.isalnum():  # 数字和字母保留
                    result.append(char.lower())
            return ''.join(result)
        
        abbr = get_pinyin_abbr(media_name)
        vertical_image = f"http://36.133.168.235:18181/img/{abbr}_st.jpg"
        horizontal_image = f"http://36.133.168.235:18181/img/{abbr}_ht.jpg"
        
        dynamic_props = {
            '作者列表': author_list,
            '清晰度': 1,
            '语言': language,
            '主演': actors,
            '内容类型': content_type,
            '上映年份': release_year,
            '关键字': keywords,
            '评分': rating,
            '推荐语': recommendation,
            '总集数': total_episodes,
            '产品分类': product_category,
            '竖图': vertical_image,
            '描述': description,
            '横图': horizontal_image,
            '版权': 1,
            '二级分类': secondary_category
        }
        
        drama_insert = "INSERT INTO drama_main (customer_id, drama_name, dynamic_properties) VALUES (%s, %s, %s)"
        cursor.execute(drama_insert, (customer_id, media_name, json.dumps(dynamic_props, ensure_ascii=False)))
        conn.commit()
        new_drama_id = cursor.lastrowid
        
        # 2. 创建子集数据
        if total_episodes > 0:
            # 确定媒体路径中的目录
            if "儿童" in str(content_type):
                content_dir = "shaoer"
            elif "教育" in str(content_type):
                content_dir = "mqxt"
            elif "电竞" in str(content_type):
                content_dir = "rywg"
            else:
                content_dir = "shaoer"
            
            for episode_num in range(1, total_episodes + 1):
                episode_name = f"{media_name}第{episode_num:02d}集"
                media_url = f"ftp://ftpmediazjyd:rD2q0y1M5eI@36.133.168.235:2121/media/hnyd/{content_dir}/{abbr}/{abbr}{episode_num:03d}.ts"
                
                # 从 video_scan_result 表匹配时长和文件大小
                duration = 0
                file_size = 0
                
                match_query = """
                    SELECT duration_formatted, size_bytes 
                    FROM video_scan_result 
                    WHERE source_file LIKE %s AND file_name LIKE %s
                    LIMIT 1
                """
                cursor.execute(match_query, (f"%{media_name}%", f"%第{episode_num}%"))
                match_result = cursor.fetchone()
                
                if match_result:
                    duration = match_result['duration_formatted'] if match_result['duration_formatted'] else 0
                    file_size = int(match_result['size_bytes']) if match_result['size_bytes'] else 0
                
                episode_props = {
                    '媒体拉取地址': media_url,
                    '媒体类型': 1,
                    '编码格式': 1,
                    '集数': episode_num,
                    '时长': duration,
                    '文件大小': file_size
                }
                
                episode_insert = "INSERT INTO drama_episode (drama_id, episode_name, dynamic_properties) VALUES (%s, %s, %s)"
                cursor.execute(episode_insert, (new_drama_id, episode_name, json.dumps(episode_props, ensure_ascii=False)))
            
            conn.commit()
        
        # 3. 插入版权方数据（包含 drama_id 和 customer_id）
        fields = [
            'drama_id', 'customer_id', 'media_name', 'upstream_copyright', 'category_level1', 'category_level1_henan',
            'category_level2_henan', 'episode_count', 'single_episode_duration', 'total_duration',
            'production_year', 'production_region', 'language', 'language_henan', 'country',
            'director', 'screenwriter', 'cast_members', 'recommendation', 'synopsis',
            'keywords', 'video_quality', 'license_number', 'rating', 'exclusive_status',
            'copyright_start_date', 'copyright_end_date', 'authorization_region',
            'authorization_platform', 'cooperation_mode'
        ]
        
        insert_fields = ['drama_id', 'customer_id']
        insert_values = [new_drama_id, customer_id]
        
        for field in fields[2:]:  # 跳过 drama_id 和 customer_id
            if field in data and data[field] is not None:
                insert_fields.append(field)
                insert_values.append(data[field])
        
        columns = ', '.join(insert_fields)
        placeholders = ', '.join(['%s'] * len(insert_fields))
        insert_query = f"INSERT INTO copyright_content ({columns}) VALUES ({placeholders})"
        cursor.execute(insert_query, insert_values)
        conn.commit()
        new_copyright_id = cursor.lastrowid
        
        conn.close()
        
        return {
            "code": 200,
            "message": "创建成功",
            "data": {
                "copyright_id": new_copyright_id,
                "drama_id": new_drama_id,
                "episodes_created": total_episodes
            }
        }
    except HTTPException:
        if conn:
            conn.close()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/copyright/{item_id}")
async def update_copyright(item_id: int, data: Dict[str, Any] = Body(...)):
    """更新版权方数据"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 检查是否存在
        check_query = "SELECT * FROM copyright_content WHERE id = %s"
        cursor.execute(check_query, (item_id,))
        item = cursor.fetchone()
        
        if not item:
            conn.close()
            raise HTTPException(status_code=404, detail="数据不存在")
        
        fields = [
            'media_name', 'upstream_copyright', 'category_level1', 'category_level1_henan',
            'category_level2_henan', 'episode_count', 'single_episode_duration', 'total_duration',
            'production_year', 'production_region', 'language', 'language_henan', 'country',
            'director', 'screenwriter', 'cast_members', 'recommendation', 'synopsis',
            'keywords', 'video_quality', 'license_number', 'rating', 'exclusive_status',
            'copyright_start_date', 'copyright_end_date', 'authorization_region',
            'authorization_platform', 'cooperation_mode'
        ]
        
        update_parts = []
        update_values = []
        
        for field in fields:
            if field in data:
                update_parts.append(f"{field} = %s")
                update_values.append(data[field])
        
        if update_parts:
            update_query = f"UPDATE copyright_content SET {', '.join(update_parts)} WHERE id = %s"
            update_values.append(item_id)
            cursor.execute(update_query, update_values)
            conn.commit()
        
        conn.close()
        
        return {
            "code": 200,
            "message": "更新成功",
            "data": {"id": item_id}
        }
    except HTTPException:
        if conn:
            conn.close()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/copyright/{item_id}")
async def delete_copyright(item_id: int):
    """删除版权方数据及关联的剧集和子集"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 查询版权信息及关联的drama_id
        check_query = "SELECT id, drama_id FROM copyright_content WHERE id = %s"
        cursor.execute(check_query, (item_id,))
        item = cursor.fetchone()
        
        if not item:
            conn.close()
            raise HTTPException(status_code=404, detail="数据不存在")
        
        drama_id = item.get('drama_id')
        
        # 如果有关联的剧集，先删除子集，再删除剧集
        if drama_id:
            # 删除子集
            cursor.execute("DELETE FROM drama_episode WHERE drama_id = %s", (drama_id,))
            # 删除剧集
            cursor.execute("DELETE FROM drama_main WHERE drama_id = %s", (drama_id,))
        
        # 删除版权信息
        delete_query = "DELETE FROM copyright_content WHERE id = %s"
        cursor.execute(delete_query, (item_id,))
        conn.commit()
        
        conn.close()
        
        return {
            "code": 200,
            "message": "删除成功",
            "data": {"id": item_id, "drama_id": drama_id}
        }
    except HTTPException:
        if conn:
            conn.close()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)


# ==================== 子集（Episode）API ====================

@app.post("/api/dramas/{drama_id}/episodes")
async def create_episode(drama_id: int, episode_data: Dict[str, Any] = Body(...)):
    """创建子集"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 检查剧集是否存在
        check_query = "SELECT * FROM drama_main WHERE drama_id = %s"
        cursor.execute(check_query, (drama_id,))
        drama = cursor.fetchone()
        
        if not drama:
            conn.close()
            raise HTTPException(status_code=404, detail="剧集不存在")
        
        # 检查必填字段
        if '节目名称' not in episode_data or not episode_data['节目名称']:
            conn.close()
            raise HTTPException(status_code=400, detail="节目名称不能为空")
        
        episode_name = episode_data['节目名称']
        
        # 构建dynamic_properties
        dynamic_props = {}
        dynamic_fields = ['媒体拉取地址', '媒体类型', '编码格式', '集数', '时长', '文件大小']
        
        for field in dynamic_fields:
            if field in episode_data:
                dynamic_props[field] = episode_data[field]
        
        # 插入数据
        insert_query = "INSERT INTO drama_episode (drama_id, episode_name, dynamic_properties) VALUES (%s, %s, %s)"
        dynamic_props_json = json.dumps(dynamic_props, ensure_ascii=False) if dynamic_props else None
        cursor.execute(insert_query, (drama_id, episode_name, dynamic_props_json))
        conn.commit()
        
        new_episode_id = cursor.lastrowid
        conn.close()
        
        return {
            "code": 200,
            "message": "创建成功",
            "data": {"episode_id": new_episode_id}
        }
    except HTTPException:
        if conn:
            conn.close()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/dramas/{drama_id}/episodes/{episode_id}")
async def update_episode(drama_id: int, episode_id: int, episode_data: Dict[str, Any] = Body(...)):
    """更新子集"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 检查子集是否存在
        check_query = "SELECT * FROM drama_episode WHERE episode_id = %s AND drama_id = %s"
        cursor.execute(check_query, (episode_id, drama_id))
        episode = cursor.fetchone()
        
        if not episode:
            conn.close()
            raise HTTPException(status_code=404, detail="子集不存在")
        
        # 构建dynamic_properties
        dynamic_props = {}
        dynamic_fields = ['媒体拉取地址', '媒体类型', '编码格式', '集数', '时长', '文件大小']
        
        for field in dynamic_fields:
            if field in episode_data:
                dynamic_props[field] = episode_data[field]
        
        # 更新数据
        update_fields = []
        update_values = []
        
        if '节目名称' in episode_data:
            update_fields.append("episode_name = %s")
            update_values.append(episode_data['节目名称'])
        
        if dynamic_props:
            update_fields.append("dynamic_properties = %s")
            update_values.append(json.dumps(dynamic_props, ensure_ascii=False))
        
        if update_fields:
            update_query = f"UPDATE drama_episode SET {', '.join(update_fields)} WHERE episode_id = %s"
            update_values.append(episode_id)
            cursor.execute(update_query, update_values)
            conn.commit()
        
        conn.close()
        
        return {
            "code": 200,
            "message": "更新成功",
            "data": {"episode_id": episode_id}
        }
    except HTTPException:
        if conn:
            conn.close()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/dramas/{drama_id}/episodes/{episode_id}")
async def delete_episode(drama_id: int, episode_id: int):
    """删除子集"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # 检查子集是否存在
        check_query = "SELECT * FROM drama_episode WHERE episode_id = %s AND drama_id = %s"
        cursor.execute(check_query, (episode_id, drama_id))
        episode = cursor.fetchone()
        
        if not episode:
            conn.close()
            raise HTTPException(status_code=404, detail="子集不存在")
        
        # 删除子集
        delete_query = "DELETE FROM drama_episode WHERE episode_id = %s"
        cursor.execute(delete_query, (episode_id,))
        conn.commit()
        
        conn.close()
        
        return {
            "code": 200,
            "message": "删除成功",
            "data": {"episode_id": episode_id}
        }
    except HTTPException:
        if conn:
            conn.close()
        raise
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))
