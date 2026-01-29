"""
Excel导出服务
提供剧集数据的Excel导出功能，支持多客户格式
"""
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from config import CUSTOMER_CONFIGS
from services.drama_service import (
    JIANGSU_HEADERS, JIANGSU_COL_WIDTHS,
    build_drama_display_dict_fast, build_episode_display_dict_fast,
    build_picture_data_fast, get_column_names, preprocess_dramas,
    preprocess_episodes, group_episodes_by_drama
)


class ExcelExportService:
    """Excel导出服务"""
    
    @staticmethod
    def format_excel_sheets(writer, customer_code: str):
        """格式化Excel表格 - 固定列宽，不换行，长文本截断显示"""
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.row_dimensions[1].height = 20
            
            for col_idx in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col_idx)
                header_cell = sheet.cell(row=1, column=col_idx)
                header_name = str(header_cell.value) if header_cell.value else ''
                
                # 根据表头长度计算列宽，限制最大30
                header_width = sum(2 if ord(c) > 127 else 1 for c in header_name)
                col_width = min(header_width + 4, 30)
                sheet.column_dimensions[col_letter].width = col_width
                
                # 设置所有单元格：不换行，垂直居中
                for row_idx in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=False, vertical='center')
    
    @staticmethod
    def format_jiangsu_excel(writer):
        """为江苏新媒体格式化Excel
        格式：第1行英文字段名，第2行中文说明，第3行开始是数据
        """
        workbook = writer.book
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        header_alignment = Alignment(wrap_text=False, vertical='center', horizontal='center')
        
        for sheet_name in workbook.sheetnames:
            if sheet_name not in JIANGSU_HEADERS:
                continue
                
            sheet = workbook[sheet_name]
            config = JIANGSU_HEADERS[sheet_name]
            
            # 插入一行作为第二行（中文说明行）
            sheet.insert_rows(2)
            
            # 设置行高
            sheet.row_dimensions[1].height = 20
            sheet.row_dimensions[2].height = 20
            
            # 处理表头和列宽
            for col_idx, field_name in enumerate(config['row1'], 1):
                col_letter = get_column_letter(col_idx)
                
                # 设置列宽
                sheet.column_dimensions[col_letter].width = JIANGSU_COL_WIDTHS.get(field_name, 15)
                
                # 第1行：英文字段名
                cell1 = sheet.cell(row=1, column=col_idx)
                cell1.value = field_name
                cell1.alignment = header_alignment
                cell1.fill = header_fill
                cell1.border = thin_border
                
                # 第2行：中文说明
                cell2 = sheet.cell(row=2, column=col_idx)
                cell2.value = config['row2'][col_idx - 1]
                cell2.alignment = header_alignment
                cell2.fill = header_fill
                cell2.border = thin_border
    
    @staticmethod
    def write_jiangsu_sheet_fast(writer, sheet_name: str, df: pd.DataFrame, header_format):
        """使用xlsxwriter快速写入江苏新媒体表"""
        headers = JIANGSU_HEADERS.get(sheet_name)
        if headers is None:
            return
        df_for_sheet = df.reindex(columns=headers['row1']) if df is not None else pd.DataFrame(columns=headers['row1'])
        df_for_sheet.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.write_row(0, 0, headers['row1'], header_format)
        worksheet.write_row(1, 0, headers['row2'], header_format)
        worksheet.freeze_panes(2, 0)
        for col_idx, field_name in enumerate(headers['row1']):
            worksheet.set_column(col_idx, col_idx, JIANGSU_COL_WIDTHS.get(field_name, 15))
    
    @staticmethod
    def build_jiangsu_excel_fast(drama_df: pd.DataFrame, episode_df: pd.DataFrame, picture_df: pd.DataFrame) -> BytesIO:
        """构建江苏新媒体Excel（xlsxwriter写入，性能更好）"""
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book
            header_format = workbook.add_format({
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#E0E0E0',
                'border': 1,
            })
            ExcelExportService.write_jiangsu_sheet_fast(writer, '剧头', drama_df, header_format)
            ExcelExportService.write_jiangsu_sheet_fast(writer, '子集', episode_df, header_format)
            ExcelExportService.write_jiangsu_sheet_fast(writer, '图片', picture_df, header_format)
        output.seek(0)
        return output
    
    @staticmethod
    def export_customer_dramas(dramas: list, episodes: list, customer_code: str) -> BytesIO:
        """导出指定客户的剧集数据为Excel
        
        Args:
            dramas: 剧集列表（已预处理）
            episodes: 子集列表（已预处理）
            customer_code: 客户代码
        
        Returns:
            BytesIO: Excel文件流
        """
        config = CUSTOMER_CONFIGS.get(customer_code, {})
        drama_columns = get_column_names(customer_code, 'drama')
        episode_columns = get_column_names(customer_code, 'episode')
        drama_col_configs = config.get('drama_columns', [])
        episode_col_configs = config.get('episode_columns', [])
        
        # 按 drama_id 分组子集
        episodes_by_drama = group_episodes_by_drama(episodes)
        
        # 构建数据
        drama_list = []
        all_episodes = []
        all_pictures = []
        
        drama_sequence = 0
        episode_sequence = 0
        picture_sequence = 0
        
        for drama in dramas:
            drama_sequence += 1
            header_dict = build_drama_display_dict_fast(drama, customer_code, drama_col_configs)
            
            # 处理序号字段
            first_col = drama_columns[0] if drama_columns else None
            if first_col and ('vod_no' in first_col.lower() or first_col == '序号'):
                header_dict[first_col] = drama_sequence
            elif first_col:
                header_dict[first_col] = ''
            
            # 江苏新媒体: sId字段留空
            if customer_code == 'jiangsu_newmedia' and 'sId' in header_dict:
                header_dict['sId'] = None
            
            drama_list.append(header_dict)
            
            # 获取子集
            drama_episodes = episodes_by_drama.get(drama['drama_id'], [])
            drama_name = drama.get('drama_name', '')
            
            for episode in drama_episodes:
                episode_sequence += 1
                ep_data = build_episode_display_dict_fast(episode, customer_code, episode_col_configs, drama_name)
                
                # 处理序号字段
                first_ep_col = episode_columns[0] if episode_columns else None
                if first_ep_col:
                    if 'vod_info_no' in first_ep_col.lower() or first_ep_col == '序号':
                        ep_data[first_ep_col] = episode_sequence
                    else:
                        ep_data[first_ep_col] = ''
                
                # 处理剧头序号关联
                if 'vod_no' in episode_columns:
                    ep_data['vod_no'] = drama_sequence
                
                # 江苏新媒体: sId和pId字段留空
                if customer_code == 'jiangsu_newmedia':
                    if 'sId' in ep_data:
                        ep_data['sId'] = None
                    if 'pId' in ep_data:
                        ep_data['pId'] = None
                
                all_episodes.append(ep_data)
            
            # 江苏新媒体的图片数据
            if customer_code == 'jiangsu_newmedia':
                abbr = drama.get('_pinyin_abbr', '')
                for pic in build_picture_data_fast(abbr):
                    picture_sequence += 1
                    pic['picture_no'] = picture_sequence
                    pic['vod_no'] = drama_sequence
                    all_pictures.append(pic)
        
        # 创建DataFrame
        drama_df = pd.DataFrame(drama_list, columns=drama_columns)
        episode_df = pd.DataFrame(all_episodes, columns=episode_columns)
        
        # 生成Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            drama_df.to_excel(writer, sheet_name='剧头', index=False)
            episode_df.to_excel(writer, sheet_name='子集', index=False)
            
            if customer_code == 'jiangsu_newmedia' and all_pictures:
                picture_columns = [col['col'] for col in config.get('picture_columns', [])]
                picture_df = pd.DataFrame(all_pictures, columns=picture_columns)
                picture_df.to_excel(writer, sheet_name='图片', index=False)
            
            if customer_code == 'jiangsu_newmedia':
                ExcelExportService.format_jiangsu_excel(writer)
            else:
                ExcelExportService.format_excel_sheets(writer, customer_code)
        
        output.seek(0)
        return output
    
    @staticmethod
    def export_single_drama(drama: dict, episodes: list, customer_code: str) -> BytesIO:
        """导出单个剧集为Excel"""
        config = CUSTOMER_CONFIGS.get(customer_code, {})
        drama_columns = get_column_names(customer_code, 'drama')
        episode_columns = get_column_names(customer_code, 'episode')
        drama_col_configs = config.get('drama_columns', [])
        episode_col_configs = config.get('episode_columns', [])
        
        # 预处理
        drama['_parsed_props'] = drama.get('_parsed_props') or {}
        drama['_pinyin_abbr'] = drama.get('_pinyin_abbr', '')
        
        # 构建剧头数据
        header_dict = build_drama_display_dict_fast(drama, customer_code, drama_col_configs)
        
        # 江苏新媒体：设置序号为1，sId留空
        if customer_code == 'jiangsu_newmedia':
            header_dict['vod_no'] = 1
            header_dict['sId'] = None
        elif drama_columns and drama_columns[0] in header_dict:
            header_dict[drama_columns[0]] = ''
        
        header_df = pd.DataFrame([header_dict], columns=drama_columns)
        
        # 构建子集数据
        episode_list = []
        drama_name = drama.get('drama_name', '')
        for i, episode in enumerate(episodes, 1):
            ep_data = build_episode_display_dict_fast(episode, customer_code, episode_col_configs, drama_name)
            
            if customer_code == 'jiangsu_newmedia':
                ep_data['vod_info_no'] = i
                ep_data['vod_no'] = 1
                ep_data['sId'] = None
                ep_data['pId'] = None
            elif episode_columns and episode_columns[0] in ep_data:
                ep_data[episode_columns[0]] = ''
            
            episode_list.append(ep_data)
        
        subset_df = pd.DataFrame(episode_list, columns=episode_columns)
        
        # 生成Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            header_df.to_excel(writer, sheet_name='剧头', index=False)
            subset_df.to_excel(writer, sheet_name='子集', index=False)
            
            if customer_code == 'jiangsu_newmedia':
                picture_data = build_picture_data_fast(drama.get('_pinyin_abbr', ''))
                for i, pic in enumerate(picture_data, 1):
                    pic['picture_no'] = i
                    pic['vod_no'] = 1
                picture_columns = [col['col'] for col in config.get('picture_columns', [])]
                picture_df = pd.DataFrame(picture_data, columns=picture_columns)
                picture_df.to_excel(writer, sheet_name='图片', index=False)
            
            if customer_code == 'jiangsu_newmedia':
                ExcelExportService.format_jiangsu_excel(writer)
            else:
                ExcelExportService.format_excel_sheets(writer, customer_code)
        
        output.seek(0)
        return output
